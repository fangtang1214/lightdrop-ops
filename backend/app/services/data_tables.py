from __future__ import annotations

import hashlib
import json
import re
import shutil
from filecmp import cmp as same_file_content
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook

from backend.app.models.schemas import (
    BadProductRecord,
    ConversionDropRanking,
    ConversionDropRankRecord,
    GoodProductRanking,
    GoodProductRankRecord,
    LiveProductRankRecord,
    TableProcessSession,
    TableProcessSummary,
)
from backend.app.services.scanner import SUPPORTED_IMAGE_EXTENSIONS, read_image_metadata

DATA_TABLES_DIR_NAME = "每场直播数据表格"
ALL_IMAGES_DIR_NAME = "所有品图片数据库"
BAD_IMAGES_DIR_NAME = "每场差品图片"

BOTTOM_BAD_PRODUCT_COUNT = 20
BadProductRule = Literal["bottom_20", "zero_conversion"]
DEFAULT_BAD_PRODUCT_RULE: BadProductRule = "bottom_20"
GOOD_PRODUCT_MERGE_RANGES = (4, 10)
REPO_ROOT = Path(__file__).resolve().parents[3]
PRODUCT_IMAGE_INDEX_CACHE_PATH = REPO_ROOT / "data" / "cache" / "product-image-index-cache.json"


@dataclass(frozen=True)
class DatabaseDirs:
    root: Path
    data_tables: Path
    all_images: Path
    bad_images: Path


def resolve_database_dirs(root_path: str) -> DatabaseDirs | None:
    root = Path(root_path).expanduser().resolve()
    data_tables = root / DATA_TABLES_DIR_NAME
    all_images = root / ALL_IMAGES_DIR_NAME
    bad_images = root / BAD_IMAGES_DIR_NAME

    if data_tables.is_dir() and all_images.is_dir() and bad_images.is_dir():
        return DatabaseDirs(
            root=root,
            data_tables=data_tables,
            all_images=all_images,
            bad_images=bad_images,
        )
    return None


def process_live_data_tables(
    root_path: str,
    *,
    bad_product_rule: BadProductRule = DEFAULT_BAD_PRODUCT_RULE,
    include_good_rankings: bool = True,
    include_conversion_drop_rankings: bool = True,
) -> TableProcessSummary | None:
    dirs = resolve_database_dirs(root_path)
    if dirs is None:
        return None

    source_signature = build_table_process_signature(dirs)

    table_files = sorted(
        [
            path
            for path in dirs.data_tables.iterdir()
            if path.is_file()
            and path.suffix.lower() in {".xlsx", ".xlsm"}
            and not path.name.startswith("~$")
        ],
        key=lambda item: item.name.lower(),
    )

    if not table_files:
        return TableProcessSummary(
            enabled=True,
            data_tables_dir=str(dirs.data_tables),
            all_images_dir=str(dirs.all_images),
            bad_images_dir=str(dirs.bad_images),
            bad_product_rule=bad_product_rule,
            source_signature=source_signature,
            message="没有找到直播数据表格",
        )

    image_index = build_product_image_index(dirs.all_images)
    sessions = [
        process_table_file(
            table_file,
            dirs.bad_images,
            image_index,
            bad_product_rule=bad_product_rule,
        )
        for table_file in table_files
    ]
    good_rankings = build_good_rankings(sessions) if include_good_rankings else []
    conversion_drop_rankings = (
        build_conversion_drop_rankings(sessions)
        if include_conversion_drop_rankings
        else []
    )

    return TableProcessSummary(
        enabled=True,
        data_tables_dir=str(dirs.data_tables),
        all_images_dir=str(dirs.all_images),
        bad_images_dir=str(dirs.bad_images),
        bad_product_rule=bad_product_rule,
        processed_sessions=sessions,
        good_rankings=good_rankings,
        conversion_drop_rankings=conversion_drop_rankings,
        source_signature=source_signature,
    )


def build_table_process_signature(dirs: DatabaseDirs) -> str:
    payload = {
        "tables": directory_file_signature(dirs.data_tables, {".xlsx", ".xlsm"}),
        "images": directory_file_signature(dirs.all_images, SUPPORTED_IMAGE_EXTENSIONS),
    }
    return hashlib.sha1(
        json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()


def current_table_process_signature(root_path: str) -> str | None:
    dirs = resolve_database_dirs(root_path)
    if dirs is None:
        return None
    return build_table_process_signature(dirs)


def directory_file_signature(root: Path, suffixes: set[str]) -> list[tuple[str, int, int]]:
    if not root.exists():
        return []

    items: list[tuple[str, int, int]] = []
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in suffixes:
            continue
        try:
            stat = path.stat()
        except OSError:
            continue
        items.append((
            str(path.relative_to(root)).replace("\\", "/").lower(),
            stat.st_size,
            stat.st_mtime_ns,
        ))
    return sorted(items)


def process_table_file(
    table_file: Path,
    bad_images_root: Path,
    image_index: dict[str, list[Path]],
    *,
    bad_product_rule: BadProductRule = DEFAULT_BAD_PRODUCT_RULE,
) -> TableProcessSession:
    products = read_conversion_rows(table_file)
    selected_bad_products = select_bad_products(products, bad_product_rule)
    bad_product_ids = {product.product_id for product in selected_bad_products}
    ranked_products = build_ranked_products(products, bad_product_ids, image_index)

    session_name = build_session_name(table_file)
    session_dir = bad_images_root / session_name
    prepare_session_output_dir(session_dir)

    output_products: list[BadProductRecord] = []
    for index, product in enumerate(selected_bad_products, start=1):
        matched_images = find_images_for_product(product.product_id, image_index)
        copied_images = copy_product_images(
            matched_images,
            session_dir,
            rank=index,
            product_id=product.product_id,
        )
        output_products.append(
            BadProductRecord(
                product_id=product.product_id,
                title=product.title,
                rank_from_bottom=index,
                exposure_click_rate=product.exposure_click_rate,
                click_conversion_rate=product.click_conversion_rate,
                refund_rate=product.refund_rate,
                net_conversion_score=product.net_conversion_score,
                exposure_people=product.exposure_people,
                click_people=product.click_people,
                deal_people=product.deal_people,
                deal_orders=product.deal_orders,
                refund_orders=product.refund_orders,
                matched_images=[str(path) for path in matched_images],
                copied_images=[str(path) for path in copied_images],
            )
        )

    unmatched = [
        product.product_id
        for product in output_products
        if not product.matched_images
    ]

    return TableProcessSession(
        table_file=str(table_file),
        session_name=session_name,
        total_products=len(products),
        selected_bad_products=len(output_products),
        matched_products=len([product for product in output_products if product.matched_images]),
        copied_images=sum(len(product.copied_images) for product in output_products),
        unmatched_product_ids=unmatched,
        products=output_products,
        ranked_products=ranked_products,
    )


def select_bad_products(
    products: list["ConversionProduct"],
    bad_product_rule: BadProductRule,
) -> list["ConversionProduct"]:
    products_by_low_score = sorted(
        products,
        key=lambda item: (
            item.net_conversion_score,
            item.click_conversion_rate,
            item.exposure_click_rate,
            -item.refund_rate,
            item.product_id,
        ),
    )

    if bad_product_rule == "zero_conversion":
        return [
            product
            for product in products_by_low_score
            if product.click_conversion_rate == 0
        ]

    return products_by_low_score[:BOTTOM_BAD_PRODUCT_COUNT]


def prepare_session_output_dir(session_dir: Path) -> None:
    session_dir.mkdir(parents=True, exist_ok=True)
    for child in session_dir.iterdir():
        if child.is_file() and re.match(r"^\d{2,}_", child.name):
            child.unlink()


@dataclass(frozen=True)
class ConversionProduct:
    product_id: str
    title: str
    exposure_click_rate: float
    click_conversion_rate: float
    refund_rate: float
    net_conversion_score: float
    exposure_people: int
    click_people: int
    deal_people: int
    deal_orders: int
    refund_orders: int


def build_ranked_products(
    products: list[ConversionProduct],
    bad_product_ids: set[str],
    image_index: dict[str, list[Path]],
) -> list[LiveProductRankRecord]:
    ranked = sorted(
        products,
        key=lambda item: (
            -item.click_conversion_rate,
            -item.deal_people,
            -item.deal_orders,
            -item.net_conversion_score,
            -item.exposure_click_rate,
            item.refund_rate,
            item.product_id,
        ),
    )
    total = len(ranked)
    records: list[LiveProductRankRecord] = []
    for index, product in enumerate(ranked, start=1):
        matched_images = find_images_for_product(product.product_id, image_index)
        records.append(
            LiveProductRankRecord(
                product_id=product.product_id,
                title=product.title,
                representative_image=str(matched_images[0]) if matched_images else None,
                matched_images=[str(path) for path in matched_images],
                rank=index,
                rank_from_bottom=total - index + 1,
                is_bad_product=product.product_id in bad_product_ids,
                exposure_click_rate=product.exposure_click_rate,
                click_conversion_rate=product.click_conversion_rate,
                refund_rate=product.refund_rate,
                net_conversion_score=product.net_conversion_score,
                exposure_people=product.exposure_people,
                click_people=product.click_people,
                deal_people=product.deal_people,
                deal_orders=product.deal_orders,
                refund_orders=product.refund_orders,
            )
        )
    return records


def build_good_rankings(sessions: list[TableProcessSession]) -> list[GoodProductRanking]:
    rankings: list[GoodProductRanking] = []
    image_hash_cache: dict[str, str] = {}

    for range_size in GOOD_PRODUCT_MERGE_RANGES:
        selected_sessions = sessions[-range_size:]
        if len(selected_sessions) < range_size:
            rankings.append(
                GoodProductRanking(
                    range_size=range_size,
                    session_names=[session.session_name for session in selected_sessions],
                    products=[],
                )
            )
            continue

        products = merge_good_products(selected_sessions, image_hash_cache)
        rankings.append(
            GoodProductRanking(
                range_size=range_size,
                session_names=[session.session_name for session in selected_sessions],
                products=products,
            )
        )

    return rankings


def build_conversion_drop_rankings(
    sessions: list[TableProcessSession],
) -> list[ConversionDropRanking]:
    rankings: list[ConversionDropRanking] = []
    image_hash_cache: dict[str, str] = {}

    for session_index, session in enumerate(sessions):
        history_sessions = sessions[:session_index]
        products = build_session_conversion_drop_ranking(
            session,
            history_sessions,
            image_hash_cache,
        )
        rankings.append(
            ConversionDropRanking(
                session_name=session.session_name,
                products=products,
            )
        )

    return rankings


def build_session_conversion_drop_ranking(
    current_session: TableProcessSession,
    history_sessions: list[TableProcessSession],
    image_hash_cache: dict[str, str],
) -> list[ConversionDropRankRecord]:
    history_by_identity: dict[str, list[tuple[TableProcessSession, LiveProductRankRecord]]] = {}
    for history_session in history_sessions:
        for history_product in history_session.ranked_products:
            for identity in build_good_product_identity_keys(history_product, image_hash_cache):
                history_by_identity.setdefault(identity, []).append((history_session, history_product))

    drop_records: list[ConversionDropRankRecord] = []
    for current_product in current_session.ranked_products:
        identities = build_good_product_identity_keys(current_product, image_hash_cache)
        candidates: dict[tuple[str, str], tuple[TableProcessSession, LiveProductRankRecord]] = {}
        for identity in identities:
            for history_session, history_product in history_by_identity.get(identity, []):
                candidates[(history_session.session_name, history_product.product_id)] = (
                    history_session,
                    history_product,
                )

        if not candidates:
            continue

        history_session, history_product = max(
            candidates.values(),
            key=lambda item: (
                item[1].click_conversion_rate,
                item[1].deal_people,
                item[1].deal_orders,
                item[1].net_conversion_score,
                item[1].exposure_click_rate,
                -item[1].refund_rate,
            ),
        )
        conversion_rate_drop = round(
            history_product.click_conversion_rate - current_product.click_conversion_rate,
            6,
        )
        net_drop = round(
            history_product.net_conversion_score - current_product.net_conversion_score,
            8,
        )
        if conversion_rate_drop <= 0:
            continue

        drop_records.append(
            ConversionDropRankRecord(
                product_id=current_product.product_id,
                title=current_product.title,
                representative_image=current_product.representative_image,
                matched_images=current_product.matched_images,
                rank=0,
                current_session_name=current_session.session_name,
                history_session_name=history_session.session_name,
                history_product_id=history_product.product_id,
                history_title=history_product.title,
                current_rank=current_product.rank,
                history_rank=history_product.rank,
                current_net_conversion_score=current_product.net_conversion_score,
                history_net_conversion_score=history_product.net_conversion_score,
                net_conversion_drop=net_drop,
                drop_ratio=round(
                    safe_rate(conversion_rate_drop, history_product.click_conversion_rate),
                    6,
                ),
                current_click_conversion_rate=current_product.click_conversion_rate,
                history_click_conversion_rate=history_product.click_conversion_rate,
                click_conversion_rate_drop=conversion_rate_drop,
                current_exposure_click_rate=current_product.exposure_click_rate,
                history_exposure_click_rate=history_product.exposure_click_rate,
                current_exposure_people=current_product.exposure_people,
                history_exposure_people=history_product.exposure_people,
                current_click_people=current_product.click_people,
                history_click_people=history_product.click_people,
                current_deal_people=current_product.deal_people,
                history_deal_people=history_product.deal_people,
                source_product_ids=sorted(
                    {current_product.product_id, history_product.product_id}
                ),
            )
        )

    ranked_records = sorted(
        drop_records,
        key=lambda item: (
            -item.click_conversion_rate_drop,
            -item.drop_ratio,
            -item.history_click_conversion_rate,
            item.current_click_conversion_rate,
            -item.net_conversion_drop,
            item.product_id,
        ),
    )
    for rank, record in enumerate(ranked_records, start=1):
        record.rank = rank
    return ranked_records


class GoodProductAccumulator:
    def __init__(self) -> None:
        self.product_id = ""
        self.title = ""
        self.representative_image: str | None = None
        self.exposure_people = 0
        self.click_people = 0
        self.deal_people = 0
        self.deal_orders = 0
        self.refund_orders = 0
        self.source_sessions: set[str] = set()
        self.source_product_ids: set[str] = set()
        self.matched_images: set[str] = set()
        self.latest_session_index = 0
        self.latest_session_name = ""
        self.identity_keys: set[str] = set()


def merge_good_products(
    sessions: list[TableProcessSession],
    image_hash_cache: dict[str, str],
) -> list[GoodProductRankRecord]:
    groups: list[GoodProductAccumulator | None] = []
    identity_to_group: dict[str, int] = {}

    for session_index, session in enumerate(sessions, start=1):
        for product in session.ranked_products:
            identities = build_good_product_identity_keys(product, image_hash_cache)
            existing_group_indices = {
                identity_to_group[identity]
                for identity in identities
                if identity in identity_to_group and groups[identity_to_group[identity]] is not None
            }

            if existing_group_indices:
                group_index = min(existing_group_indices)
                group = groups[group_index]
                if group is None:
                    group = GoodProductAccumulator()
                    groups[group_index] = group
                for other_index in sorted(existing_group_indices - {group_index}):
                    other_group = groups[other_index]
                    if other_group is None:
                        continue
                    merge_good_product_accumulators(group, other_group)
                    for identity in other_group.identity_keys:
                        identity_to_group[identity] = group_index
                    groups[other_index] = None
            else:
                group_index = len(groups)
                group = GoodProductAccumulator()
                groups.append(group)

            add_product_to_good_group(group, product, session.session_name, session_index, identities)
            for identity in group.identity_keys:
                identity_to_group[identity] = group_index

    ranked_groups = sorted(
        [group for group in groups if group is not None],
        key=lambda group: (
            -safe_rate(group.deal_people, group.click_people),
            -group.deal_people,
            -group.deal_orders,
            -build_net_conversion_score(group),
            -safe_rate(group.click_people, group.exposure_people),
            safe_rate(group.refund_orders, group.deal_orders),
            group.product_id,
        ),
    )

    return [
        build_good_product_record(group, rank=index)
        for index, group in enumerate(ranked_groups, start=1)
    ]


def build_good_product_identity_keys(
    product: LiveProductRankRecord,
    image_hash_cache: dict[str, str],
) -> set[str]:
    identities: set[str] = set()
    for image in product.matched_images:
        fingerprint = image_fingerprint(image, image_hash_cache)
        if fingerprint:
            identities.add(f"image:{fingerprint}")

    if not identities:
        identities.add(f"product:{product.product_id}")
    return identities


def image_fingerprint(image_path: str, image_hash_cache: dict[str, str]) -> str:
    if image_path not in image_hash_cache:
        try:
            _, _, _, image_hash, _ = read_image_metadata(Path(image_path))
            image_hash_cache[image_path] = image_hash
        except Exception:
            image_hash_cache[image_path] = str(Path(image_path).resolve()).lower()
    return image_hash_cache[image_path]


def add_product_to_good_group(
    group: GoodProductAccumulator,
    product: LiveProductRankRecord,
    session_name: str,
    session_index: int,
    identities: set[str],
) -> None:
    group.exposure_people += product.exposure_people
    group.click_people += product.click_people
    group.deal_people += product.deal_people
    group.deal_orders += product.deal_orders
    group.refund_orders += product.refund_orders
    group.source_sessions.add(session_name)
    group.source_product_ids.add(product.product_id)
    group.matched_images.update(product.matched_images)
    group.identity_keys.update(identities)

    if session_index >= group.latest_session_index:
        group.product_id = product.product_id
        group.title = product.title
        group.representative_image = product.representative_image or group.representative_image
        group.latest_session_index = session_index
        group.latest_session_name = session_name


def merge_good_product_accumulators(
    target: GoodProductAccumulator,
    source: GoodProductAccumulator,
) -> None:
    target.exposure_people += source.exposure_people
    target.click_people += source.click_people
    target.deal_people += source.deal_people
    target.deal_orders += source.deal_orders
    target.refund_orders += source.refund_orders
    target.source_sessions.update(source.source_sessions)
    target.source_product_ids.update(source.source_product_ids)
    target.matched_images.update(source.matched_images)
    target.identity_keys.update(source.identity_keys)

    if source.latest_session_index >= target.latest_session_index:
        target.product_id = source.product_id
        target.title = source.title
        target.representative_image = source.representative_image or target.representative_image
        target.latest_session_index = source.latest_session_index
        target.latest_session_name = source.latest_session_name


def build_good_product_record(
    group: GoodProductAccumulator,
    rank: int,
) -> GoodProductRankRecord:
    exposure_click_rate = round(safe_rate(group.click_people, group.exposure_people), 6)
    click_conversion_rate = round(safe_rate(group.deal_people, group.click_people), 6)
    refund_rate = round(safe_rate(group.refund_orders, group.deal_orders), 6)
    net_conversion_score = round(build_net_conversion_score(group), 8)

    return GoodProductRankRecord(
        product_id=group.product_id,
        title=group.title,
        representative_image=group.representative_image,
        matched_images=sorted(group.matched_images),
        rank=rank,
        exposure_click_rate=exposure_click_rate,
        click_conversion_rate=click_conversion_rate,
        refund_rate=refund_rate,
        net_conversion_score=net_conversion_score,
        exposure_people=group.exposure_people,
        click_people=group.click_people,
        deal_people=group.deal_people,
        deal_orders=group.deal_orders,
        refund_orders=group.refund_orders,
        source_session_count=len(group.source_sessions),
        source_sessions=sorted(group.source_sessions),
        source_product_ids=sorted(group.source_product_ids),
        latest_session_name=group.latest_session_name,
    )


def build_net_conversion_score(group: GoodProductAccumulator) -> float:
    return (
        safe_rate(group.click_people, group.exposure_people)
        * safe_rate(group.deal_people, group.click_people)
        * (1 - safe_rate(group.refund_orders, group.deal_orders))
    )


def read_conversion_rows(table_file: Path) -> list[ConversionProduct]:
    workbook = load_workbook(table_file, data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = {str(value).strip(): index for index, value in enumerate(header_row) if value}

        required_headers = ["商品id", "商品标题"]
        missing = [header for header in required_headers if header not in headers]
        if missing:
            raise ValueError(f"{table_file.name} 缺少字段：{', '.join(missing)}")

        products: list[ConversionProduct] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            product_id = get_text(row, headers, "商品id")
            if not product_id:
                continue

            exposure_people = get_int(row, headers, "曝光人数")
            click_people = get_int(row, headers, "点击人数")
            deal_people = get_int(row, headers, "成交人数")
            deal_orders = get_int(row, headers, "成交订单数")
            refund_orders = get_int(row, headers, "退款订单数")

            exposure_click_rate = safe_rate(click_people, exposure_people)
            click_conversion_rate = safe_rate(deal_people, click_people)
            refund_rate = safe_rate(refund_orders, deal_orders)
            net_conversion_score = (
                exposure_click_rate * click_conversion_rate * (1 - refund_rate)
            )

            products.append(
                ConversionProduct(
                    product_id=product_id,
                    title=get_text(row, headers, "商品标题"),
                    exposure_click_rate=round(exposure_click_rate, 6),
                    click_conversion_rate=round(click_conversion_rate, 6),
                    refund_rate=round(refund_rate, 6),
                    net_conversion_score=round(net_conversion_score, 8),
                    exposure_people=exposure_people,
                    click_people=click_people,
                    deal_people=deal_people,
                    deal_orders=deal_orders,
                    refund_orders=refund_orders,
                )
            )

        return products
    finally:
        workbook.close()


def build_product_image_index(all_images_root: Path) -> dict[str, list[Path]]:
    signature = build_product_image_index_signature(all_images_root)
    cached = load_product_image_index_cache(all_images_root, signature)
    if cached is not None:
        return cached

    index: dict[str, list[Path]] = {}
    image_paths = [
        path
        for path in all_images_root.rglob("*")
        if path.is_file() and path.suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS
    ]

    for path in image_paths:
        source_text = str(path.relative_to(all_images_root))
        for product_id in set(re.findall(r"\d{8,}", source_text)):
            index.setdefault(product_id, []).append(path)

    save_product_image_index_cache(all_images_root, signature, index)
    return index


def build_product_image_index_signature(all_images_root: Path) -> str:
    payload = directory_file_signature(all_images_root, SUPPORTED_IMAGE_EXTENSIONS)
    return hashlib.sha1(
        json.dumps(payload, ensure_ascii=False).encode("utf-8")
    ).hexdigest()


def load_product_image_index_cache(
    all_images_root: Path,
    signature: str,
) -> dict[str, list[Path]] | None:
    if not PRODUCT_IMAGE_INDEX_CACHE_PATH.exists():
        return None

    try:
        payload = json.loads(PRODUCT_IMAGE_INDEX_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None

    root_key = str(all_images_root.resolve())
    entry = payload.get(root_key)
    if not isinstance(entry, dict) or entry.get("signature") != signature:
        return None

    raw_index = entry.get("index")
    if not isinstance(raw_index, dict):
        return None

    return {
        str(product_id): [Path(path) for path in paths if isinstance(path, str)]
        for product_id, paths in raw_index.items()
        if isinstance(paths, list)
    }


def save_product_image_index_cache(
    all_images_root: Path,
    signature: str,
    index: dict[str, list[Path]],
) -> None:
    PRODUCT_IMAGE_INDEX_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    try:
        payload = json.loads(PRODUCT_IMAGE_INDEX_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        payload = {}

    root_key = str(all_images_root.resolve())
    payload[root_key] = {
        "signature": signature,
        "index": {
            product_id: [str(path) for path in paths]
            for product_id, paths in index.items()
        },
    }
    PRODUCT_IMAGE_INDEX_CACHE_PATH.write_text(
        json.dumps(payload, ensure_ascii=False),
        encoding="utf-8",
    )


def find_images_for_product(product_id: str, image_index: dict[str, list[Path]]) -> list[Path]:
    return sorted(image_index.get(product_id, []), key=lambda item: str(item).lower())


def copy_product_images(
    images: list[Path],
    session_dir: Path,
    rank: int,
    product_id: str,
) -> list[Path]:
    copied: list[Path] = []
    for image in images:
        destination = reusable_destination(
            image,
            session_dir / f"{rank:02d}_{product_id}_{image.name}",
        )
        if not destination.exists():
            shutil.copy2(image, destination)
        copied.append(destination)
    return copied


def reusable_destination(source: Path, destination: Path) -> Path:
    if not destination.exists():
        return destination
    try:
        if source.stat().st_size == destination.stat().st_size and same_file_content(
            source,
            destination,
            shallow=False,
        ):
            return destination
    except OSError:
        pass
    return unique_destination(destination)


def unique_destination(destination: Path) -> Path:
    if not destination.exists():
        return destination

    stem = destination.stem
    suffix = destination.suffix
    for index in range(2, 1000):
        candidate = destination.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"无法生成唯一文件名：{destination}")


def build_session_name(table_file: Path) -> str:
    name = table_file.stem
    for suffix in ("直播间商品列表", "商品列表", "直播数据", "数据表格"):
        name = name.replace(suffix, "")
    return name.strip(" _-") or table_file.stem


def get_text(row: tuple[Any, ...], headers: dict[str, int], name: str) -> str:
    index = headers.get(name)
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def get_int(row: tuple[Any, ...], headers: dict[str, int], name: str) -> int:
    value = get_value(row, headers, name)
    if value in (None, ""):
        return 0
    if isinstance(value, str):
        value = value.replace(",", "").strip()
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def get_value(row: tuple[Any, ...], headers: dict[str, int], name: str) -> Any:
    index = headers.get(name)
    if index is None or index >= len(row):
        return None
    return row[index]


def safe_rate(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0
    return max(numerator / denominator, 0)
